"""
1. Aktüel ürünler sayfasına git
    https://www.bim.com.tr/Categories/100/aktuel-urunler.aspx
2. Gelecek Hafta'ya tıkla
3. Tarihlere tıklayarak aşağıdaki işlemleri yap
    3.1. Ürünler içinde aşağıdaki işlemleri
        3.1.1. Ürün resmini kaydet
        3.1.2. Ürün bilgilerini oku
        3.1.3. Ürün fiyatını oku
"""
from moduller.tarayici import Tarayici
from selenium.webdriver.common.by import By
from time import sleep
from urllib import request
from openpyxl import Workbook
from openpyxl import load_workbook
import os

# Excel dosyası yoksa oluştur. Varsa aç
excel_yolu = "./bim.xlsx"
if os.path.exists(excel_yolu):
    ck = load_workbook(excel_yolu)
    cs = ck.active
else:
    ck = Workbook()
    cs = ck.active
    cs.append(["Ad", "Görsel", "Marka", "Açıklama", "Fiyat"])

# Tarayıcıyı oluştur
tarayici_nesne = Tarayici()
tarayici = tarayici_nesne.al()

# 1. Aktüel ürünler sayfasına git
tarayici.get("https://www.bim.com.tr/Categories/100/aktuel-urunler.aspx")
sleep(2)

# Çerezleri kabul et
# tarayici.find_element(By.XPATH, '//*[@id="form1"]/div/footer/div/div[5]/button[1]').click()
tarayici.find_element(By.XPATH, "//button[contains(text(), 'Kabul Et')]").click()

# 2. Gelecek Hafta'ya tıkla
tarayici.find_element(By.XPATH, "//span[normalize-space()='GELECEK HAFTA']").click()

# 3. Tarihlere tıklayarak aşağıdaki işlemleri yap
tarihler = tarayici.find_elements(By.XPATH, "//div[@class='subButtonArea subButtonArea-5 active']//a")

# tarihlerdeki işlemleri for içinde yapacağız
for i, tarih in enumerate(tarihler):
    # tarayici.execute_script("window.scrollTo(0, 0);")  yukarı kaydırma hatayı gidermedi

    # sayfa yenilenince elemanlar tekrar oluşturuluyor.
    # Bu sebeple döngü içinde sıradaki elemanı yeniden seçmemiz gerekiyor.
    tarih = tarayici.find_element(By.XPATH, f"//div[@class='subButtonArea subButtonArea-5 active']//a[{i+1}]")

    # ekran görüntüleri alındığında sayfa aşağı iniyor ve tarih tıklamada üst çubuk engel oluyor. Javascript ile tıklama yapacağız
    tarayici.execute_script("arguments[0].click();", tarih)
    # tarih.click()  # sayfa yenileniyor
    sleep(2)

    # Daha fazla ürün göster tuşuna; kaybolana kadar tıklama işlemi
    while True:
        try:
            tarayici.find_element(By.XPATH, "//a[@href='javascript:;changeLPage();']").click()
        except:
            break

    # 3.1. Ürünler içinde aşağıdaki işlemleri
    urunler = tarayici.find_elements(By.XPATH, "//div[contains(@class, 'product')]")
    for urun in urunler:
        try:
            # ürünün adını alalım
            ad = urun.find_element(By.XPATH, ".//h2[@class='title']")
        except:
            # ürün adını alırken hata alırsa; ürün adı yoktur demektir
            continue

        try:
            # 3.1.1. Ürün resmini kaydet
            # eski: urun.find_element(By.TAG_NAME, "img").screenshot(f"./gorseller/{ad.text}.png")
            img = urun.find_element(By.TAG_NAME, "img")
            img_src = img.get_attribute("src")
            img_adi = img_src.split("/")[-1]
            img_yolu = f"./gorseller/{img_adi}"
            request.urlretrieve(img_src, img_yolu)
        except:
            # görünmeyen ürünlerin ekran görüntüsü alınırken hata alırsa geç
            continue

        # 3.1.2. Ürün bilgilerini oku
        # ürün markasını varsa alalım
        try:
            marka = urun.find_element(By.XPATH, ".//h2[@class='subTitle']").text
        except:
            marka = ""

        # ürünün açıklaması varsa alalım
        try:
            aciklama = urun.find_element(By.XPATH, ".//div[@class='textArea']").text
        except:
            aciklama = ""

        # 3.1.3. Ürün fiyatını oku
        fiyat = urun.find_element(By.XPATH, ".//a[@class='gButton triangle']").text

        cs.append([
            ad.text,
            f'=HYPERLINK("{os.getcwd()}\\{img_yolu}", "GÖRSEL")',
            marka,
            aciklama,
            fiyat
        ])


tarayici.quit()
ck.save(excel_yolu)