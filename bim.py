from moduller.tarayici import Tarayici
from selenium.webdriver.common.by import By
from time import sleep
from urllib import request
from openpyxl import Workbook
from   openpyxl import  load_workbook
import os


excel_yolu =  ".bim.xlsx"
if os.path.exists(excel_yolu):
    ck = load_workbook(excel_yolu)
    cs = ck.active
else:
    ck = Workbook()
    cs = ck.active
    cs.append(["Ad", "Görsel", "Marka", "Açıklama", "Fiyat"])
tarayici_nesne = Tarayici()
tarayici = tarayici_nesne.al()

tarayici.get("https://www.bim.com.tr/Categories/100/aktuel-urunler.aspx")


tarayici.find_element(By.XPATH,"//button[contains(text(), 'Kabul Et')]").click()


tarayici.find_element(By.XPATH, "//span[normalize-space()='GELECEK HAFTA']").click()

tarihler = tarayici.find_elements(By.XPATH, "//div[@class='subButtonArea subButtonArea-5 active']//a")


for i, tarih in enumerate(tarihler):
    tarih = tarayici.find_element(By.XPATH, f"//div[@class='subButtonArea subButtonArea-5 active']//a[{i + 1}]")
    tarayici.execute_script("arguments[0].click();", tarih)
    sleep(2)

    while True:
        try:
            tarayici.find_element(By.XPATH, "//a[href='javascript:;changeLPage();']").click()
        except:
             break


    urunler = tarayici.find_elements(By.XPATH, "//div[contains(@class, 'product')]")
    for urun in urunler:
        try:
            ad = urun.find_element(By.XPATH, ".//h2[@class='title']")
        except:
            continue

        try:
            img = urun.find_element(By.TAG_NAME, "img")
            img_src = img.get_attribute("src")
            img_adi = img_src.split("/")[-1]
            request.urlretrieve(img_src, f"./gorseller/{img_adi}")
        except:

            continue

        try:
           marka = urun.find_element(By.XPATH, ".//h2[@class='subTitle']").text
        except:
           marka = ""

        try:
            aciklama = urun.find_element(By.XPATH, ".div[@class='textArea']").text
        except:
            aciklama = ""


        fiyat = urun.find_element(By.XPATH, ".//a[@class='gButton triangle']").text

        print("-"*50)
        print("Ad:", ad.text)
        print("Marka:", marka)
        print("Açıklama:", aciklama)
        print("Fiyat:", fiyat.replace("\n", ""))

        tarayici.quit()
        ck.save(excel_yolu)





